from pathlib import Path
from sys import exit, path

from openpyxl import load_workbook

# Add all the paths in environment
folder_path = str(Path(__file__).parent.absolute())
package_path = str(Path(folder_path).parent.absolute())
file_path = str(Path(__file__).absolute())

path.append(file_path)
path.append(folder_path)
path.append(package_path)

# Import later to avoid import error
import FlameSoft.fs as fs


def string_to_list(var: str):
    """Convert the excel input to list"""
    try:
        ans = [int(num) for num in var.split(',')]
    except Exception as _:
        ans = []
    return ans


def read_txt(path: str):
    """Function to read the points.txt and output list of points"""
    ans = []
    with open(path, 'r') as file:
        for line in file:
            ans.append(line.strip())
    ans = [int(i) for i in ans]

    x1, y1, x2, y2 = ans
    ans = [(x1, y1), (x2, y2)]
    return ans


path_to_excel = folder_path + r'\\FlameSoft.xlsm'

try:
    wb = load_workbook(path_to_excel, read_only=True, data_only=True)
    ws = wb['Control_Sheet']

    vals = dict(out=ws['F6'].value,
                path=ws['F7'].value,
                slices=ws['F8'].value,
                filter=string_to_list(ws['F9'].value),
                thresh=string_to_list(ws['F10'].value),
                flow=ws['F11'].value,
                operation=ws['F12'].value,
                length=ws['F13'].value,
                fps=ws['F14'].value,
                crop_previous=ws['F15'].value,
                hratio=ws['F16'].value,
                sub_frame=ws['F17'].value)
    print(vals)
    cls = fs.Flame(path_=vals['path'], out=vals['out'])

except Exception as _:
    print("Error:  ", _)

try:
    if vals['operation'] == 1:
        # Check if crop from previous is enabled
        if vals['crop_previous']:

            try:

                points = read_txt(vals['out'] + r'\bin' + r'\points.txt')

            except Exception as _:

                print("Please recerop the image(error in reading points.txt")
        else:

            points = fs.Crop(path_=vals['path'], out=vals['out']).crop_video()

        cls.process(breaks=vals['slices'], filter_size=vals['filter'], thresh_val=vals['thresh'], crop_points=points,
                    flow_right=vals['flow'], height=vals['hratio'], sub_frame=vals['sub_frame'])

    elif vals['operation'] == 2:
        cls.whiten_image()

    elif vals['operation'] == 3:
        cls.blacken_image()

    elif vals['operation'] == 4:
        cls.get_data()

    elif vals['operation'] == 5:
        cls.view_pimage()

    elif vals['operation'] == 6:
        cls.Dataframe(length=vals['length'], fps=vals['fps'])

    exit()

except Exception as _:
    print("Error:  ", _)
    exit()
